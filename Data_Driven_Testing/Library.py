import json
import jsonpath
import requests
import openpyxl


class Common:

    def __init__(self, Filepath, sheetname):
        global wk
        global sh
        wk = openpyxl.load_workbook(Filepath)
        sh = wk[sheetname]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        cols = sh.max_column
        return cols

    def fetch_Column_name(self):
        column_count = self.fetch_column_count()
        li = []
        for i in range(1, column_count + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i - 1, cell.value)
        return li

    def update_Request_with_data(self, rownumber, jsonrequest, Columnlist):
        cols = self.fetch_column_count()
        for i in range(1, cols + 1):
            cell = sh.cell(row=rownumber, column=i)
            jsonrequest[Columnlist[i - 1]] = cell.value
        return jsonrequest
