import requests
import json
import jsonpath
import pytest
import openpyxl

@pytest.fixture(scope = "module")
def fixture_code():
    print("------------------------------------------")
    print("This is Starting of Code")
    print("------------------------------------------")
    yield
    print("------------------------------------------")
    print("This is End of Code")
    print("------------------------------------------")

def test_login(fixture_code):
    #API
    API_url = "https://reqres.in/api/login"
    Request_file = "C://Users//shaz5//Desktop//Learn//APITesting//PytestLearning//Files//Login.json"
    Requestobj = open(Request_file, "r")
    jsonrequest = json.loads(Requestobj.read())
    #Excel
    wk = openpyxl.load_workbook("C://Users//shaz5//Desktop//Learn//APITesting//PytestLearning//Files//loginopenpyxl.xlsx")
    sh= wk['Sheet1']
    rows =sh.max_row
    for i in range(2,rows+1):
        cell_email = sh.cell(row=i,column=1)
        cell_password = sh.cell(row=i,column=2)
        jsonrequest['email'] = cell_email.value
        jsonrequest['password'] = cell_password.value
        print(jsonrequest)
        Response = requests.post(API_url, jsonrequest)
        print(Response.text)
        print(Response.status_code)
    #jsonresponse = json.loads(Response.text)
    #First_Name = jsonpath.jsonpath(jsonresponse,"data.first_name")
    #id = jsonpath.jsonpath(jsonresponse,"data.id")
    #print(First_Name)
    #assert Response.status_code == 200